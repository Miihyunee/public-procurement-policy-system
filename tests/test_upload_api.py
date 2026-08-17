"""
tests.test_upload_api

**업로드 API** 검증 — 양식 다운로드 · 파일 검증 · 저장.

여기서 가장 중요하게 고정하는 성질:

1. 오류가 하나라도 있으면 **DB 에 아무것도 저장되지 않는다** (전부 검증 → 전부 저장)
2. 검증 API 는 **저장하지 않는다** — 저장은 별도 엔드포인트의 책임이다
3. 오류는 **행 번호 · 항목명 · 내용**으로 돌아온다
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from procurement.app import create_app
from procurement.database.bootstrap import bootstrap
from procurement.database.purchase_repository import PurchaseRepository
from procurement.uploads.format import header_row
from procurement.uploads.template import DATA_SHEET_TITLE, write_template

VALIDATE_URL = "/uploads/purchases/validate"
GOOD_ROW: list[object] = [
    date(2026, 3, 15),
    date(2026, 2, 20),
    date(2026, 4, 1),
    "한빛산업개발",
    "220-81-62517",
    54648000,
]


@pytest.fixture
def db_path(tmp_path: Path) -> Path:
    """초기화된 빈 DB."""
    path = tmp_path / "upload.db"
    bootstrap(path)
    return path


@pytest.fixture
def client(db_path: Path) -> TestClient:
    """격리 DB 를 쓰는 API 클라이언트."""
    return TestClient(create_app(db_path=db_path))


def _excel(path: Path, rows: list[list[object]]) -> Path:
    """표준 머리글 + 주어진 행으로 엑셀을 만듭니다."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(list(header_row()))
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


class TestTemplateDownload:
    """``GET /uploads/template``."""

    def test_returns_xlsx(self, client: TestClient) -> None:
        response = client.get("/uploads/template")

        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        assert response.content[:2] == b"PK"  # xlsx 는 zip 컨테이너

    def test_has_attachment_filename(self, client: TestClient) -> None:
        response = client.get("/uploads/template")

        disposition = response.headers["content-disposition"]
        assert disposition.startswith("attachment")
        assert "UTF-8''" in disposition  # 한글 파일명

    def test_downloaded_file_is_the_standard_form(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        """내려받은 파일이 표준 양식 그대로다."""
        path = tmp_path / "downloaded.xlsx"
        path.write_bytes(client.get("/uploads/template").content)

        workbook = load_workbook(path)
        try:
            sheet = workbook[DATA_SHEET_TITLE]
            assert tuple(cell.value for cell in sheet[1]) == header_row()
        finally:
            workbook.close()


class TestValidateEndpoint:
    """``POST /uploads/purchases/validate``."""

    def test_valid_file_passes(self, client: TestClient, tmp_path: Path) -> None:
        path = _excel(tmp_path / "good.xlsx", [GOOD_ROW])

        body = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()

        assert body["ok"] is True
        assert body["total_rows"] == 1
        assert body["valid_rows"] == 1
        assert body["error_rows"] == 0
        assert body["issues"] == []

    def test_row_errors_are_reported_with_row_and_header(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        """오류는 **행 번호 · 항목명 · 내용**으로 돌아온다."""
        bad = ["2026.13.45", date(2026, 2, 20), date(2026, 4, 1), None, "999", "abc"]
        path = _excel(tmp_path / "bad.xlsx", [GOOD_ROW, bad])

        body = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()

        assert body["ok"] is False
        assert body["total_rows"] == 2
        assert body["valid_rows"] == 1
        assert body["error_rows"] == 1

        headers = {issue["header"] for issue in body["issues"]}
        assert {"결의일자", "기업명", "사업자등록번호", "계"} <= headers
        assert all(issue["row_number"] == 3 for issue in body["issues"])

    def test_missing_header_is_a_file_error(self, client: TestClient, tmp_path: Path) -> None:
        """항목이 빠지면 행 오류가 아니라 **파일 오류**로 보고한다."""
        path = tmp_path / "partial.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.append(["결의일자", "계약일자"])
        sheet.append([date(2026, 3, 15), date(2026, 2, 20)])
        workbook.save(path)
        workbook.close()

        body = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()

        assert body["ok"] is False
        assert body["file_errors"]
        assert "기업명" in body["file_errors"][0]

    def test_unreadable_file_is_reported_not_raised(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        """읽을 수 없는 파일도 500 이 아니라 결과로 돌려준다."""
        path = tmp_path / "broken.xlsx"
        path.write_bytes(b"not an excel file")

        response = client.post(VALIDATE_URL, json={"file_path": str(path)})

        assert response.status_code == 200
        assert response.json()["file_errors"]

    def test_wrong_extension_is_rejected(self, client: TestClient, tmp_path: Path) -> None:
        path = tmp_path / "data.csv"
        path.write_text("a,b", encoding="utf-8")

        body = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()

        assert ".xlsx" in body["file_errors"][0]

    def test_missing_file_path_is_422(self, client: TestClient) -> None:
        assert client.post(VALIDATE_URL, json={}).status_code == 422

    def test_empty_file_path_is_422(self, client: TestClient) -> None:
        assert client.post(VALIDATE_URL, json={"file_path": ""}).status_code == 422

    def test_template_itself_validates(self, client: TestClient, tmp_path: Path) -> None:
        """내려받은 양식의 예시 행이 검증을 통과한다(양식과 규칙이 일치)."""
        path = write_template(tmp_path / "t.xlsx")

        body = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()

        assert body["ok"] is True, body["issues"]


class TestValidateEndpointNeverStores:
    """⛔ 검증 API 는 어떤 경우에도 저장하지 않는다."""

    def test_valid_file_does_not_store(
        self, client: TestClient, db_path: Path, tmp_path: Path
    ) -> None:
        """정상 파일이어도 검증 API 는 **저장하지 않는다**(책임 분리)."""
        path = _excel(tmp_path / "good.xlsx", [GOOD_ROW])

        body = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()

        assert body["stored"] is False
        assert body["storage_note"]
        assert PurchaseRepository(db_path).count() == 0

    def test_error_file_does_not_store(
        self, client: TestClient, db_path: Path, tmp_path: Path
    ) -> None:
        """오류가 하나라도 있으면 **정상 행도** 저장되지 않는다."""
        bad = ["2026.13.45", date(2026, 2, 20), date(2026, 4, 1), None, "999", "abc"]
        path = _excel(tmp_path / "mixed.xlsx", [GOOD_ROW, bad])

        client.post(VALIDATE_URL, json={"file_path": str(path)})

        assert PurchaseRepository(db_path).count() == 0

    def test_storage_note_says_validation_only(
        self, client: TestClient, tmp_path: Path
    ) -> None:
        """검증 전용 API 임을 사용자에게 문장으로 알려 준다.

        .. note::
            **기대값이 바뀐 이유** — 이전에는 "지급일 항목이 없어 저장할 수
            없다" 가 이유였습니다. 2026-08-17 PM 결정으로 그 Blocker 가
            해소되어 저장은 별도 엔드포인트(``POST /uploads/purchases``)가
            담당하며, 이 API 는 **검증 전용**으로 남았습니다.
        """
        path = _excel(tmp_path / "good.xlsx", [GOOD_ROW])

        note = client.post(VALIDATE_URL, json={"file_path": str(path)}).json()["storage_note"]

        assert "저장하지 않았습니다" in note
