"""
tests.test_upload_template

**표준 양식 ``.xlsx`` 생성** 검증.

가장 중요한 성질은 두 가지입니다.

1. 양식이 **확정된 컬럼만** 담는다 (미확정 컬럼이 새어 나가지 않는다)
2. 양식으로 만든 파일을 **우리 어댑터가 그대로 읽을 수 있다** (왕복 성립)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from procurement.uploads.excel_adapter import read_standard_workbook
from procurement.uploads.format import PENDING_COLUMNS, example_row, header_row
from procurement.uploads.template import (
    DATA_SHEET_TITLE,
    GUIDE_SHEET_TITLE,
    TEMPLATE_FILE_NAME,
    build_template_bytes,
    build_template_workbook,
    write_template,
)


class TestStructure:
    """시트 구성."""

    def test_has_data_and_guide_sheets(self) -> None:
        workbook = build_template_workbook()
        try:
            assert workbook.sheetnames == [DATA_SHEET_TITLE, GUIDE_SHEET_TITLE]
        finally:
            workbook.close()

    def test_first_sheet_is_the_data_sheet(self) -> None:
        """어댑터는 **첫 시트**를 읽으므로 데이터 시트가 앞에 있어야 한다."""
        workbook = build_template_workbook()
        try:
            assert workbook.sheetnames[0] == DATA_SHEET_TITLE
        finally:
            workbook.close()

    def test_header_row_matches_the_definition(self) -> None:
        workbook = build_template_workbook()
        try:
            sheet = workbook[DATA_SHEET_TITLE]
            values = tuple(cell.value for cell in sheet[1])
            assert values == header_row()
        finally:
            workbook.close()

    def test_example_row_is_present(self) -> None:
        workbook = build_template_workbook()
        try:
            sheet = workbook[DATA_SHEET_TITLE]
            values = tuple(str(cell.value) for cell in sheet[2])
            assert values == example_row()
        finally:
            workbook.close()

    def test_guide_mentions_vat(self) -> None:
        """`계` 가 VAT 포함 총액이라는 점을 양식 안에서 알려 준다."""
        workbook = build_template_workbook()
        try:
            sheet = workbook[GUIDE_SHEET_TITLE]
            text = "\n".join(str(row[0].value) for row in sheet.iter_rows(max_col=1))
            assert "부가가치세" in text
            assert "공급가액" in text
        finally:
            workbook.close()


class TestNoUnconfirmedColumns:
    """⛔ 확정되지 않은 컬럼이 양식에 들어가지 않는다."""

    def test_pending_columns_are_absent(self) -> None:
        workbook = build_template_workbook()
        try:
            sheet = workbook[DATA_SHEET_TITLE]
            headers = {cell.value for cell in sheet[1]}
            assert not headers & set(PENDING_COLUMNS)
        finally:
            workbook.close()

    def test_exactly_six_columns(self) -> None:
        """확정 컬럼은 6개다.

        .. note::
            **기대값이 바뀐 이유** — 2026-08-17 PM 결정으로 표준 양식에
            ``지급일`` 이 추가되었습니다(5 → 6). ``payment_date`` 를 nullable
            로 바꾸는 대신 양식에서 받기로 한 결정입니다.
        """
        assert len(header_row()) == 6


class TestRoundTrip:
    """양식 → 파일 → 어댑터 읽기가 성립한다."""

    def test_written_file_is_readable(self, tmp_path: Path) -> None:
        path = write_template(tmp_path / TEMPLATE_FILE_NAME)

        result = read_standard_workbook(path)

        assert result.headers == header_row()
        assert result.sheet_name == DATA_SHEET_TITLE
        # 예시 행 한 줄이 들어 있다.
        assert result.row_count == 1

    def test_bytes_and_file_are_equivalent(self, tmp_path: Path) -> None:
        path = tmp_path / "t.xlsx"
        path.write_bytes(build_template_bytes())

        workbook = load_workbook(path)
        try:
            assert workbook.sheetnames[0] == DATA_SHEET_TITLE
        finally:
            workbook.close()

    def test_creates_parent_directory(self, tmp_path: Path) -> None:
        path = write_template(tmp_path / "nested" / "deep" / "t.xlsx")
        assert path.exists()

    def test_file_name_is_xlsx(self) -> None:
        assert TEMPLATE_FILE_NAME.endswith(".xlsx")
