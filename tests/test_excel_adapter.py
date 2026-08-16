"""
tests.test_excel_adapter

**표준 양식 ``.xlsx`` 읽기 어댑터** 검증.

이 어댑터는 **읽기만** 합니다. 값의 업무적 타당성은 검증 계층이 판정하므로,
여기서는 "엑셀에서 무엇이 어떻게 나오는가" 만 고정합니다.

.. note::
    ``openpyxl`` 의존성은 2026-08-16 지시서 §11 에서 승인되었습니다.
"""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook

from procurement.uploads.excel_adapter import (
    MAX_FILE_BYTES,
    ExcelReadError,
    read_standard_workbook,
)
from procurement.uploads.format import header_row

HEADERS = list(header_row())
GOOD_ROW = [date(2026, 3, 15), date(2026, 2, 20), "한빛산업개발", "220-81-62517", 54648000]


def _write(path: Path, rows: list[list[object]], *, headers: list[object] | None = None) -> Path:
    """시트 하나짜리 엑셀 파일을 만듭니다."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "구매실적"
    sheet.append(HEADERS if headers is None else headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


class TestNormalRead:
    """정상 파일을 읽는다."""

    def test_reads_headers_and_rows(self, tmp_path: Path) -> None:
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [GOOD_ROW]))

        assert result.headers == tuple(HEADERS)
        assert result.row_count == 1
        assert result.sheet_name == "구매실적"

    def test_row_is_keyed_by_header(self, tmp_path: Path) -> None:
        """행은 **머리글 이름**을 키로 돌려준다(검증 계층이 그대로 받는다)."""
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [GOOD_ROW]))

        assert result.rows[0]["기업명"] == "한빛산업개발"
        assert result.rows[0]["계"] == 54648000

    def test_first_row_number_is_two(self, tmp_path: Path) -> None:
        """머리글이 1행이므로 첫 데이터 행은 2행이다(오류 메시지 번호와 일치)."""
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [GOOD_ROW]))
        assert result.first_row_number == 2

    def test_only_first_sheet_is_read(self, tmp_path: Path) -> None:
        """안내 시트가 뒤에 있어도 데이터 시트만 읽는다."""
        path = tmp_path / "a.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = "구매실적"
        sheet.append(HEADERS)
        sheet.append(GOOD_ROW)
        guide = workbook.create_sheet("작성안내")
        guide.append(["■ 작성 안내"])
        workbook.save(path)
        workbook.close()

        result = read_standard_workbook(path)
        assert result.sheet_name == "구매실적"
        assert result.row_count == 1


class TestValueHandling:
    """값 유형별 처리."""

    def test_datetime_becomes_date(self, tmp_path: Path) -> None:
        """엑셀 날짜는 시각이 붙어 오더라도 날짜로 정리한다."""
        row = [datetime(2026, 3, 15, 9, 30), date(2026, 2, 20), "A", "2208162517", 1000]
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [row]))

        assert result.rows[0]["결의일자"] == date(2026, 3, 15)

    def test_float_becomes_decimal(self, tmp_path: Path) -> None:
        """실수는 부동소수 오차를 피하려고 Decimal 로 바꾼다."""
        row = [date(2026, 3, 15), date(2026, 2, 20), "A", "2208162517", 1000.5]
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [row]))

        assert result.rows[0]["계"] == Decimal("1000.5")

    def test_integer_is_kept(self, tmp_path: Path) -> None:
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [GOOD_ROW]))
        assert result.rows[0]["계"] == 54648000

    def test_text_is_trimmed(self, tmp_path: Path) -> None:
        row = [date(2026, 3, 15), date(2026, 2, 20), "  A기업  ", "2208162517", 1000]
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [row]))

        assert result.rows[0]["기업명"] == "A기업"

    def test_blank_cell_becomes_none(self, tmp_path: Path) -> None:
        """빈 셀은 ``None`` 으로 **그대로** 넘긴다(임의 대체 금지)."""
        row = [date(2026, 3, 15), date(2026, 2, 20), "   ", "2208162517", 1000]
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [row]))

        assert result.rows[0]["기업명"] is None

    def test_missing_trailing_cells_are_filled_with_none(self, tmp_path: Path) -> None:
        """뒤쪽 셀이 생략된 행도 모든 머리글 키를 갖는다."""
        row = [date(2026, 3, 15), date(2026, 2, 20), "A기업"]
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", [row]))

        assert set(result.rows[0]) == set(HEADERS)
        assert result.rows[0]["계"] is None

    def test_fully_empty_rows_are_skipped(self, tmp_path: Path) -> None:
        """완전히 빈 행은 건너뛴다(엑셀 편집 흔적)."""
        result = read_standard_workbook(
            _write(tmp_path / "a.xlsx", [GOOD_ROW, [None, None, None, None, None], GOOD_ROW])
        )
        assert result.row_count == 2

    def test_partially_empty_row_is_kept(self, tmp_path: Path) -> None:
        """일부만 빈 행은 **남긴다** — 검증 계층이 오류로 보고해야 한다."""
        result = read_standard_workbook(
            _write(tmp_path / "a.xlsx", [[None, None, "A기업", None, None]])
        )
        assert result.row_count == 1


class TestFileErrors:
    """읽을 수 없는 파일은 사용자가 이해할 수 있는 문장으로 거부한다."""

    def test_wrong_extension(self, tmp_path: Path) -> None:
        path = tmp_path / "data.csv"
        path.write_text("a,b", encoding="utf-8")

        with pytest.raises(ExcelReadError, match=r"\.xlsx"):
            read_standard_workbook(path)

    def test_missing_file(self, tmp_path: Path) -> None:
        with pytest.raises(ExcelReadError, match="찾을 수 없습니다"):
            read_standard_workbook(tmp_path / "none.xlsx")

    def test_empty_file(self, tmp_path: Path) -> None:
        path = tmp_path / "empty.xlsx"
        path.write_bytes(b"")

        with pytest.raises(ExcelReadError, match="비어 있습니다"):
            read_standard_workbook(path)

    def test_corrupt_file(self, tmp_path: Path) -> None:
        path = tmp_path / "broken.xlsx"
        path.write_bytes(b"this is not a zip archive")

        with pytest.raises(ExcelReadError, match="엑셀"):
            read_standard_workbook(path)

    def test_directory_is_rejected(self, tmp_path: Path) -> None:
        path = tmp_path / "folder.xlsx"
        path.mkdir()

        with pytest.raises(ExcelReadError):
            read_standard_workbook(path)

    def test_too_large_file(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
        """지나치게 큰 파일은 읽기 전에 거부한다."""
        monkeypatch.setattr("procurement.uploads.excel_adapter.MAX_FILE_BYTES", 10)
        path = _write(tmp_path / "a.xlsx", [GOOD_ROW])

        with pytest.raises(ExcelReadError, match="너무 큽니다"):
            read_standard_workbook(path)

    def test_default_limit_is_generous(self) -> None:
        """기본 한도는 실무 파일을 막지 않을 만큼 크다."""
        assert MAX_FILE_BYTES >= 10 * 1024 * 1024

    def test_header_only_file_is_not_an_error(self, tmp_path: Path) -> None:
        """머리글만 있는 파일은 읽기는 성공하고 행이 0건이다."""
        result = read_standard_workbook(_write(tmp_path / "a.xlsx", []))
        assert result.row_count == 0

    def test_file_without_any_standard_header(self, tmp_path: Path) -> None:
        """표준 항목이 하나도 없으면 표준 양식이 아니라고 알린다."""
        path = _write(tmp_path / "a.xlsx", [["x"]], headers=["엉뚱한열"])

        with pytest.raises(ExcelReadError, match="표준 업로드 양식이 아닙니다"):
            read_standard_workbook(path)

    def test_partial_headers_are_left_to_validation(self, tmp_path: Path) -> None:
        """⛔ 일부 항목 누락은 **읽기 단계에서 판정하지 않는다.**

        어느 항목이 없는지 알려 주는 일은 검증 계층의 몫이며, 여기서 중복
        판정하면 메시지가 두 곳에서 나오게 됩니다.
        """
        path = _write(tmp_path / "a.xlsx", [["2026-03-15"]], headers=["결의일자"])

        result = read_standard_workbook(path)
        assert result.headers == ("결의일자",)
