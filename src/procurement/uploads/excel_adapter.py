"""
procurement.uploads.excel_adapter

**표준 업로드 양식 ``.xlsx`` 파일을 읽는 어댑터**입니다.

::

    .xlsx  →  [ 이 모듈 ]  →  머리글 + 행(머리글→값 매핑)  →  validation  →  ...

이 모듈이 하는 일은 **읽기뿐**입니다. 값의 업무적 타당성(날짜 형식·금액·
사업자등록번호 등)은 :mod:`procurement.uploads.validation` 이 판정합니다.
검증 규칙을 여기에 중복 구현하지 않습니다.

.. note::
    엑셀 XML 파서를 직접 만들지 않고 ``openpyxl`` 을 사용합니다(PM 승인 —
    2026-08-16 지시서 §11). 실측에서 확인했듯 ``.xlsx`` 는 문자열을
    ``sharedStrings`` 또는 ``inlineStr`` 로 저장하고 날짜를 일련번호로
    보관하므로, 직접 파싱하면 파일마다 다르게 깨집니다.

.. warning::
    **업무규칙을 판단하지 않습니다.** 빈 셀은 ``None`` 으로 그대로 전달하며,
    "빈 값을 무엇으로 대체할지" 는 검증·매핑 계층의 몫입니다.
"""

from __future__ import annotations

import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from procurement.uploads.format import REQUIRED_HEADERS

#: 허용하는 확장자. 사용자가 ``.xls`` 나 ``.csv`` 를 올리면 읽기 전에 거부한다.
ALLOWED_SUFFIXES: Final[tuple[str, ...]] = (".xlsx",)

#: 머리글이 있는 행 번호(1부터). 표준 양식은 1행이 머리글이다.
HEADER_ROW_NUMBER: Final[int] = 1

#: 읽기를 허용할 최대 파일 크기(바이트). 기관 1년치 지출 데이터를 충분히 담고도
#: 남는 크기이며, 실수로 올린 거대 파일이 메모리를 잠식하는 것을 막는다.
MAX_FILE_BYTES: Final[int] = 50 * 1024 * 1024


class ExcelReadError(Exception):
    """엑셀 파일을 읽을 수 없을 때 발생합니다.

    사용자에게 그대로 보여줄 수 있는 한국어 문장을 담습니다.
    """


@dataclass(frozen=True, kw_only=True)
class WorkbookRead:
    """엑셀에서 읽어 낸 결과.

    Attributes:
        headers: 머리글 행의 값(앞뒤 공백 제거, 빈 칸 제외).
        rows: 데이터 행. 각 행은 ``머리글 → 값`` 매핑입니다.
        first_row_number: 첫 데이터 행의 엑셀 행 번호. 오류 메시지에 그대로
            쓰이므로 **사용자가 엑셀에서 보는 번호와 같아야** 합니다.
        sheet_name: 읽은 시트 이름.
    """

    headers: tuple[str, ...] = ()
    rows: list[dict[str, object]] = field(default_factory=list)
    first_row_number: int = HEADER_ROW_NUMBER + 1
    sheet_name: str = ""

    @property
    def row_count(self) -> int:
        """읽은 데이터 행 수."""
        return len(self.rows)


def read_standard_workbook(source: str | Path) -> WorkbookRead:
    """표준 양식 ``.xlsx`` 파일을 읽습니다.

    머리글은 **검증하지 않고 그대로 반환**합니다. 필수 항목 누락 판정은
    :func:`procurement.uploads.validation.validate_headers` 가 담당하므로
    여기서 중복 판정하지 않습니다. 다만 표준 머리글이 **하나도 없으면** 표준
    양식이 아니라고 보고 읽기 단계에서 거부합니다(사용자가 엉뚱한 파일을 올린
    경우를 빨리 알려 주기 위함).

    Args:
        source: 읽을 파일 경로.

    Returns:
        :class:`WorkbookRead`.

    Raises:
        ExcelReadError: 확장자·존재·크기·형식 문제로 읽을 수 없는 경우.
    """
    path = _check_file(source)

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise ExcelReadError(
            "엑셀 파일(.xlsx)로 열 수 없습니다. 표준 양식을 내려받아 "
            "엑셀에서 저장한 파일인지 확인하세요."
        ) from exc
    except (zipfile.BadZipFile, OSError, ValueError, KeyError, TypeError) as exc:
        # 손상된 zip·누락된 내부 항목 등은 라이브러리마다 예외형이 다르다.
        # .xlsx 는 zip 컨테이너라 내용이 깨지면 zipfile 예외가 먼저 나온다.
        raise ExcelReadError(f"엑셀 파일을 읽는 중 오류가 발생했습니다: {exc}") from exc

    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            raise ExcelReadError("엑셀 파일에 시트가 없습니다.")

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            raw_header = next(rows_iter)
        except StopIteration as exc:
            raise ExcelReadError(
                "엑셀 파일이 비어 있습니다. 표준 양식의 머리글 행이 필요합니다."
            ) from exc

        headers = _clean_headers(raw_header)
        if not headers:
            raise ExcelReadError(
                "첫 행에 머리글이 없습니다. 표준 양식을 내려받아 사용하세요."
            )
        if not set(headers) & set(REQUIRED_HEADERS):
            raise ExcelReadError(
                "표준 업로드 양식이 아닙니다. 첫 행에서 표준 항목을 하나도 "
                f"찾지 못했습니다(읽은 머리글: {', '.join(headers)}). "
                "표준 양식을 내려받아 사용하세요."
            )

        rows = [
            _build_row(headers, raw_row)
            for raw_row in rows_iter
            if not _is_empty_row(raw_row)
        ]
        return WorkbookRead(
            headers=headers,
            rows=rows,
            first_row_number=HEADER_ROW_NUMBER + 1,
            sheet_name=str(sheet.title),
        )
    finally:
        workbook.close()


# ----------------------------------------------------------------------
# 내부 헬퍼
# ----------------------------------------------------------------------
def _check_file(source: str | Path) -> Path:
    """파일 존재·확장자·크기를 확인합니다."""
    path = Path(source)

    if path.suffix.lower() not in ALLOWED_SUFFIXES:
        allowed = " / ".join(ALLOWED_SUFFIXES)
        raise ExcelReadError(
            f"{allowed} 파일만 올릴 수 있습니다(올린 파일: {path.name}). "
            "엑셀에서 '다른 이름으로 저장 → Excel 통합 문서(.xlsx)' 로 저장하세요."
        )
    if not path.exists():
        raise ExcelReadError(f"파일을 찾을 수 없습니다: {path}")
    if not path.is_file():
        raise ExcelReadError(f"파일이 아닙니다: {path}")

    size = path.stat().st_size
    if size == 0:
        raise ExcelReadError("파일이 비어 있습니다(0바이트).")
    if size > MAX_FILE_BYTES:
        limit_mb = MAX_FILE_BYTES // (1024 * 1024)
        raise ExcelReadError(
            f"파일이 너무 큽니다({size / 1024 / 1024:.1f}MB). "
            f"{limit_mb}MB 이하만 올릴 수 있습니다."
        )
    return path


def _clean_headers(raw_header: tuple[object, ...]) -> tuple[str, ...]:
    """머리글 행을 문자열로 정리합니다(빈 칸 제외, 앞뒤 공백 제거)."""
    return tuple(str(value).strip() for value in raw_header if _has_text(value))


def _has_text(value: object) -> bool:
    """머리글로 쓸 수 있는 값인지 판단합니다."""
    return value is not None and str(value).strip() != ""


def _is_empty_row(raw_row: tuple[object, ...]) -> bool:
    """모든 셀이 비어 있는 행인지 판단합니다.

    엑셀은 편집 흔적 때문에 빈 행을 남기는 경우가 많으므로, **완전히 빈 행만**
    건너뜁니다. 일부만 빈 행은 그대로 넘겨 검증 계층이 오류로 보고하게 합니다.
    """
    return all(not _has_text(value) for value in raw_row)


def _build_row(headers: tuple[str, ...], raw_row: tuple[object, ...]) -> dict[str, object]:
    """머리글 → 값 매핑을 만듭니다.

    머리글보다 짧은 행(뒤쪽 빈 셀이 생략된 경우)은 ``None`` 으로 채웁니다.
    """
    row: dict[str, object] = {}
    for index, header in enumerate(headers):
        value = raw_row[index] if index < len(raw_row) else None
        row[header] = _normalize_cell(value)
    return row


def _normalize_cell(value: object) -> object:
    """셀 값을 검증 계층이 다루기 쉬운 형태로 정리합니다.

    - 날짜형은 :class:`datetime.date` 로 (엑셀 일련번호는 openpyxl 이 이미 변환)
    - 실수는 :class:`decimal.Decimal` 로 (부동소수 오차 방지)
    - 문자열은 앞뒤 공백 제거, 빈 문자열은 ``None``
    - 그 밖의 값은 **그대로** 둔다(임의 해석 금지)
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date | time):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        return Decimal(str(value))
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value
