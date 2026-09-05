"""
procurement.uploads.template

**표준 업로드 양식 ``.xlsx`` 파일을 생성**합니다.

사용자는 이 파일을 내려받아 데이터를 채운 뒤 그대로 올립니다. 기관마다 다른
원본 엑셀을 시스템이 추측해 해석하는 대신, **우리가 지정한 양식**을 쓰게 하는
것이 목적입니다.

구성:

- ``구매실적`` 시트 — 1행 머리글, 2행 입력 예시
- ``작성안내`` 시트 — 항목별 설명 (:func:`~procurement.uploads.format.guide_lines`)

.. warning::
    **확정되지 않은 컬럼을 넣지 않습니다.** 양식에 칸이 있으면 사용자가 채우고,
    그 값을 시스템이 해석하게 되어 확정되지 않은 업무규칙이 생깁니다. 미확정
    항목은 :data:`~procurement.uploads.format.PENDING_COLUMNS` 에만 남겨 둡니다.

.. note::
    컬럼 정의는 :mod:`procurement.uploads.format` 하나만 참조합니다. 양식 파일과
    검증 규칙이 **같은 정의**에서 나오므로 서로 어긋날 수 없습니다.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from procurement.uploads.format import STANDARD_COLUMNS, example_row, guide_lines, header_row

#: 데이터 입력 시트 이름.
DATA_SHEET_TITLE: Final[str] = "구매실적"

#: 안내 시트 이름.
GUIDE_SHEET_TITLE: Final[str] = "작성안내"

#: 내려받을 때 사용할 기본 파일명.
TEMPLATE_FILE_NAME: Final[str] = "구매실적_표준양식.xlsx"

#: 머리글 배경색(엑셀 표준 팔레트의 옅은 회색).
_HEADER_FILL: Final[PatternFill] = PatternFill("solid", fgColor="DDEBF7")

#: 예시 행 글자색(회색) — 지우고 쓰라는 뜻을 시각적으로 전달한다.
_EXAMPLE_FONT: Final[Font] = Font(color="808080", italic=True)


def build_template_workbook() -> Workbook:
    """표준 양식 통합 문서를 만듭니다.

    Returns:
        시트가 채워진 :class:`openpyxl.Workbook`.
    """
    workbook = Workbook()
    data_sheet = workbook.active
    assert data_sheet is not None  # 새 Workbook 은 항상 활성 시트를 가진다.
    data_sheet.title = DATA_SHEET_TITLE

    headers = header_row()
    data_sheet.append(list(headers))
    data_sheet.append(list(example_row()))

    for index, column in enumerate(STANDARD_COLUMNS, start=1):
        letter = get_column_letter(index)
        cell = data_sheet.cell(row=1, column=index)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        # 머리글·예시·설명 중 가장 긴 것에 맞춰 열 너비를 잡는다(한글 폭 고려).
        width = max(len(column.header) * 2, len(column.example) + 2, 14)
        data_sheet.column_dimensions[letter].width = min(width, 30)
        data_sheet.cell(row=2, column=index).font = _EXAMPLE_FONT

    # 머리글을 고정해 행이 많아도 항목명이 보이게 한다.
    data_sheet.freeze_panes = "A2"

    guide_sheet = workbook.create_sheet(GUIDE_SHEET_TITLE)
    for line in guide_lines():
        guide_sheet.append([line])
    guide_sheet.column_dimensions["A"].width = 80
    for row in guide_sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str) and cell.value.startswith("■"):
                cell.font = Font(bold=True)

    return workbook


def build_template_bytes() -> bytes:
    """표준 양식을 ``.xlsx`` 바이트로 만듭니다(HTTP 응답용).

    Returns:
        ``.xlsx`` 파일 내용.
    """
    buffer = io.BytesIO()
    workbook = build_template_workbook()
    try:
        workbook.save(buffer)
    finally:
        workbook.close()
    return buffer.getvalue()


def write_template(destination: str | Path) -> Path:
    """표준 양식을 파일로 저장합니다(CLI·테스트용).

    Args:
        destination: 저장할 경로. 상위 디렉터리는 자동 생성합니다.

    Returns:
        저장된 파일 경로.
    """
    path = Path(destination)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(build_template_bytes())
    return path
